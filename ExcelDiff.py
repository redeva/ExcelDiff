import openpyxl as ol
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import time

# 填充颜色
yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')  # 黄
green_fill = PatternFill(patternType='solid', fgColor='90EE90')  # 淡绿色
blue_fill = PatternFill(patternType='solid', fgColor='00CCFF')

# 字体颜色
red_font = Font(color='c00000')  # 红色字体

# 获取工作簿对象实例
book1 = ol.load_workbook('8.xlsx')
# book2 = ol.load_workbook('2.xlsx')

# 获取工作表对象实例
book1_sheet1 = book1['Sheet1']
book1_sheet2 = book1['Sheet2']

# 定义两个list ，存放工作表待处理的单元格范围
book1_sheet1_rowlist = range(book1_sheet1.min_row, book1_sheet1.max_row)  # 起始行加1 ，跳过标题行
book1_sheet1_collist = range(book1_sheet1.min_column, book1_sheet1.max_column)

# 简单的循环一下，对比区域内的数值
for j in book1_sheet1_rowlist:
    for k in book1_sheet1_collist:
        value_a = book1_sheet1.cell(row=j, column=k).value
        value_b = book1_sheet2.cell(row=j, column=k).value
        if value_a != value_b:
            print('\n=============================\n')
            print(
                f'出大事了，{value_a}和{value_b}的数值不一样！它们位于第{j}行第{k}列，对应的样品编号是{book1_sheet1.cell(row=j, column=1).value}，对应测试指标是{book1_sheet1.cell(row=1, column=k).value}')
            # 设置字体颜色及填充
            book1_sheet1.cell(row=j, column=k).font = red_font
            book1_sheet1.cell(row=j, column=k).fill = green_fill

# 最后，将book1另存成一个文件
save_time = time.strftime("%Y%m%d_%H%M%S", time.localtime())
excel_name = save_time + '.xlsx'  # 由时间命名文件名，避免重复
book1.save(excel_name)

#和github进行了合并，20240818_091348